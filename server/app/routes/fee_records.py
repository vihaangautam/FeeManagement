from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from bson import ObjectId
from datetime import datetime
from typing import Optional
from collections import defaultdict
import io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule

from app.database import get_database
from app.models import FeeRecordCreate, FeeRecordResponse

router = APIRouter(prefix="/api/fees", tags=["fees"])


async def fee_doc_to_response(doc: dict) -> dict:
    db = get_database()
    student_name = ""
    batch_name = ""
    if doc.get("student_id"):
        try:
            student = await db.students.find_one({"_id": ObjectId(doc["student_id"])})
            if student:
                student_name = student["name"]
                if student.get("batch_id"):
                    try:
                        batch = await db.batches.find_one({"_id": ObjectId(student["batch_id"])})
                        if batch:
                            batch_name = batch["name"]
                    except Exception:
                        pass
        except Exception:
            pass

    date_paid = doc.get("date_paid", datetime.utcnow())
    if isinstance(date_paid, str):
        date_paid = datetime.fromisoformat(date_paid).date()
    elif isinstance(date_paid, datetime):
        date_paid = date_paid.date()

    return {
        "id": str(doc["_id"]),
        "student_id": doc.get("student_id", ""),
        "student_name": student_name,
        "batch_name": batch_name,
        "amount_paid": doc.get("amount_paid", 0),
        "date_paid": date_paid,
        "fee_month": doc.get("fee_month", 1),
        "fee_year": doc.get("fee_year", 2026),
        "note": doc.get("note", ""),
        "created_at": doc.get("created_at", datetime.utcnow()),
    }


@router.get("")
async def list_fee_records(
    month: Optional[int] = Query(None, ge=1, le=12),
    year: Optional[int] = Query(None, ge=2020, le=2100),
):
    db = get_database()
    query = {}
    if month is not None:
        query["fee_month"] = month
    if year is not None:
        query["fee_year"] = year
    records = await db.fee_records.find(query).sort("created_at", -1).to_list(1000)
    result = []
    for r in records:
        result.append(await fee_doc_to_response(r))
    return result


@router.get("/status")
async def get_fee_status(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020, le=2100),
):
    """Returns per-student fee status for a given month: paid, pending, partial."""
    db = get_database()
    students = await db.students.find().sort("name", 1).to_list(500)
    result = []
    for student in students:
        sid = str(student["_id"])
        # Get batch name
        batch_name = ""
        if student.get("batch_id"):
            try:
                batch = await db.batches.find_one({"_id": ObjectId(student["batch_id"])})
                if batch:
                    batch_name = batch["name"]
            except Exception:
                pass
        # Sum payments for this month
        payments = await db.fee_records.find({
            "student_id": sid,
            "fee_month": month,
            "fee_year": year,
        }).to_list(50)
        total_paid = sum(p.get("amount_paid", 0) for p in payments)
        expected = student.get("monthly_fee", 0)
        if expected <= 0:
            status = "no_fee"
        elif total_paid >= expected:
            status = "paid"
        elif total_paid > 0:
            status = "partial"
        else:
            status = "pending"
        result.append({
            "student_id": sid,
            "student_name": student["name"],
            "batch_id": student.get("batch_id", ""),
            "batch_name": batch_name,
            "monthly_fee": expected,
            "total_paid": total_paid,
            "status": status,
            "payments": [
                {
                    "id": str(p["_id"]),
                    "amount_paid": p["amount_paid"],
                    "date_paid": p["date_paid"].isoformat() if isinstance(p["date_paid"], (datetime,)) else str(p["date_paid"]),
                    "note": p.get("note", ""),
                }
                for p in payments
            ],
        })
    return result


@router.get("/student/{student_id}")
async def get_student_fees(student_id: str):
    db = get_database()
    records = await db.fee_records.find({"student_id": student_id}).sort(
        [("fee_year", -1), ("fee_month", -1)]
    ).to_list(200)
    result = []
    for r in records:
        result.append(await fee_doc_to_response(r))
    return result


@router.get("/export/advanced")
async def export_fees_advanced(year: int = Query(..., ge=2020, le=2100)):
    db = get_database()
    students_cur = await db.students.find().sort("name", 1).to_list(1000)
    batches_cur = await db.batches.find().to_list(100)
    batch_map = {str(b["_id"]): b for b in batches_cur}
    
    payments = await db.fee_records.find({"fee_year": year}).to_list(10000)
    student_payments = defaultdict(lambda: defaultdict(int))
    for p in payments:
        sid = str(p.get("student_id"))
        month = p.get("fee_month")
        amt = p.get("amount_paid", 0)
        if sid and month:
            student_payments[sid][month] += amt

    wb = openpyxl.Workbook()
    
    # --- Sheet 1: Fee Matrix ---
    ws1 = wb.active
    ws1.title = "Fee Matrix"
    
    headers1 = ["Student Name", "Batch", "Type", "Fee/Month"] + [datetime(2000, m, 1).strftime("%b") for m in range(1, 13)] + ["Total Paid", "Balance"]
    ws1.append(headers1)
    
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 12
    for i in range(5, 17):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 9
    ws1.column_dimensions['Q'].width = 12
    ws1.column_dimensions['R'].width = 12

    ws1.freeze_panes = 'B2'
    
    header_font = Font(bold=True)
    for cell in ws1[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    current_month = datetime.utcnow().month
    current_year = datetime.utcnow().year
    
    students_data = []
    
    row_idx = 2
    for student in students_cur:
        sid = str(student["_id"])
        b = batch_map.get(student.get("batch_id", ""), {})
        batch_name = b.get("name", "Unknown")
        batch_type = b.get("type", "Coaching")
        monthly_fee = student.get("monthly_fee", 0)
        
        student_created_at = student.get("created_at", datetime.utcnow())
        join_year = student_created_at.year
        raw_join_month = student_created_at.month
        
        # If the student has backfilled payments before their profile creation date, 
        # consider their "actual" join month as the earliest month they paid for.
        first_paid_month = min([m for m, amt in student_payments[sid].items() if amt > 0], default=12)
        join_month = min(raw_join_month, first_paid_month) if year == join_year else raw_join_month
        
        if year < join_year:
            months_active = 0
        elif year == join_year:
            if year == current_year:
                months_active = max(0, current_month - join_month + 1)
            else:
                months_active = 12 - join_month + 1
        elif year > join_year:
            if year == current_year:
                months_active = current_month
            elif year < current_year:
                months_active = 12
            else:
                months_active = 0

        row = [student["name"], batch_name, batch_type, monthly_fee]
        
        for m in range(1, 13):
            paid = student_payments[sid].get(m)
            is_active = True
            if year < join_year or (year == join_year and m < join_month):
                is_active = False
            if year > current_year or (year == current_year and m > current_month):
                is_active = False
                
            if paid is not None and paid > 0:
                row.append(paid)
            else:
                if is_active:
                    row.append(0)
                else:
                    row.append("") # blank

        row.append(f"=SUM(E{row_idx}:P{row_idx})")
        row.append(f"=D{row_idx}*{months_active}-Q{row_idx}")
        
        ws1.append(row)
        students_data.append({
            "sid": sid, 
            "batch_id": student.get("batch_id", ""), 
            "name": student["name"], 
            "phone": student.get("contact", student.get("phone", "")), 
            "join_date": student_created_at.strftime("%Y-%m-%d"), 
            "fee": monthly_fee,
            "months_active": months_active
        })
        row_idx += 1
        
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    if row_idx > 2:
        month_range = f"E2:P{row_idx-1}"
        ws1.conditional_formatting.add(month_range, CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, fill=green_fill))
        ws1.conditional_formatting.add(month_range, CellIsRule(operator='equal', formula=['0'], stopIfTrue=True, fill=red_fill))
        ws1.conditional_formatting.add(month_range, CellIsRule(operator='equal', formula=['""'], stopIfTrue=True, fill=gray_fill))
        
    # --- Sheet 2: Student List ---
    ws2 = wb.create_sheet(title="Student List")
    ws2.append(["Name", "Phone", "Batch", "Join Date", "Monthly Fee"])
    for cell in ws2[1]: cell.font = header_font
    for sd in students_data:
        bname = batch_map.get(str(sd["batch_id"]), {}).get("name", "Unknown")
        ws2.append([sd["name"], sd["phone"], bname, sd["join_date"], sd["fee"]])
        
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 15
    ws2.column_dimensions['E'].width = 15
    
    # --- Sheet 3: Summary ---
    ws3 = wb.create_sheet(title="Summary")
    ws3.append(["Batch Name"] + [datetime(2000, m, 1).strftime("%b") for m in range(1, 13)] + ["Total Collected", "Total Expected", "Outstanding Balance"])
    for cell in ws3[1]: cell.font = header_font
    ws3.column_dimensions['A'].width = 25
    for i in range(2, 14):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 9
    ws3.column_dimensions['N'].width = 15
    ws3.column_dimensions['O'].width = 15
    ws3.column_dimensions['P'].width = 18
    
    s_row_idx = 2
    for b in batches_cur:
        bid = str(b["_id"])
        bname = b["name"]
        batch_students = [sd for sd in students_data if str(sd["batch_id"]) == bid]
        
        row = [bname]
        for m in range(1, 13):
            m_total = sum(student_payments[sd["sid"]].get(m, 0) for sd in batch_students)
            row.append(m_total)
            
        row.append(f"=SUM(B{s_row_idx}:M{s_row_idx})")
        expected = sum(sd["fee"] * sd["months_active"] for sd in batch_students)
        row.append(expected)
        row.append(f"=O{s_row_idx}-N{s_row_idx}") # Expected - Collected
        ws3.append(row)
        s_row_idx += 1

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    headers = {
        'Content-Disposition': f'attachment; filename="TutorFlow_Export_{year}.xlsx"'
    }
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


@router.get("/export")
async def export_fees(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020, le=2100),
):
    """Returns structured data for Excel export of a given month."""
    db = get_database()
    students = await db.students.find().sort("name", 1).to_list(500)
    rows = []
    for student in students:
        sid = str(student["_id"])
        batch_name = ""
        if student.get("batch_id"):
            try:
                batch = await db.batches.find_one({"_id": ObjectId(student["batch_id"])})
                if batch:
                    batch_name = batch["name"]
            except Exception:
                pass
        payments = await db.fee_records.find({
            "student_id": sid,
            "fee_month": month,
            "fee_year": year,
        }).to_list(50)
        total_paid = sum(p.get("amount_paid", 0) for p in payments)
        expected = student.get("monthly_fee", 0)
        payment_dates = ", ".join(
            p["date_paid"].strftime("%d/%m/%Y") if isinstance(p["date_paid"], datetime) else str(p["date_paid"])
            for p in payments
        )
        rows.append({
            "Student Name": student["name"],
            "Batch": batch_name,
            "Phone": student.get("phone", ""),
            "Monthly Fee (₹)": expected,
            "Paid (₹)": total_paid,
            "Pending (₹)": max(0, expected - total_paid),
            "Status": "Paid" if total_paid >= expected else ("Partial" if total_paid > 0 else "Pending"),
            "Payment Date(s)": payment_dates,
        })
    return {"month": month, "year": year, "data": rows}


@router.post("", status_code=201)
async def create_fee_record(data: FeeRecordCreate):
    db = get_database()
    # Verify student exists
    try:
        student = await db.students.find_one({"_id": ObjectId(data.student_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid student ID")
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    doc = {
        "student_id": data.student_id,
        "amount_paid": data.amount_paid,
        "date_paid": datetime.combine(data.date_paid, datetime.min.time()),
        "fee_month": data.fee_month,
        "fee_year": data.fee_year,
        "note": data.note,
        "created_at": datetime.utcnow(),
    }
    result = await db.fee_records.insert_one(doc)
    doc["_id"] = result.inserted_id
    return await fee_doc_to_response(doc)


@router.delete("/{fee_id}")
async def delete_fee_record(fee_id: str):
    db = get_database()
    try:
        result = await db.fee_records.delete_one({"_id": ObjectId(fee_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid fee record ID")
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Fee record not found")
    return {"message": "Fee record deleted successfully"}
